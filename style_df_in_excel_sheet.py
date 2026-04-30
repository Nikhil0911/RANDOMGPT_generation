import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# -------- STYLES -------- #
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

fill_before = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
fill_after = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

center_align = Alignment(horizontal='center', vertical='center')

# -------- FUNCTION -------- #
def write_sheet(ws, df1, df2, sheet_name):
    ws.title = sheet_name

    # Combine side by side
    combined = pd.concat([df1, df2], axis=1)

    n1 = df1.shape[1]
    n2 = df2.shape[1]

    # Write merged headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n1)
    ws.merge_cells(start_row=1, start_column=n1+1, end_row=1, end_column=n1+n2)

    ws.cell(row=1, column=1).value = "Before"
    ws.cell(row=1, column=n1+1).value = "After"

    # Style merged headers
    for col in range(1, n1+n2+1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thick_border

    # Write dataframe
    for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)

            # Format last columns (df1 last col & df2 last col)
            if c_idx == n1 or c_idx == (n1+n2):
                if isinstance(value, (int, float)):
                    value = f"{value:,.2f}"

            cell.value = value

            # Header row styling
            if r_idx == 2:
                cell.font = bold_font
                cell.border = thick_border
            else:
                cell.border = thin_border

            # Color separation
            if c_idx <= n1:
                cell.fill = fill_before
            else:
                cell.fill = fill_after

# -------- MAIN -------- #
wb = Workbook()
wb.remove(wb.active)

# Example: dfs = [df1, df2, ..., df12]
for i in range(0, 12, 2):
    ws = wb.create_sheet(title=f"Sheet_{i//2 + 1}")
    write_sheet(ws, dfs[i], dfs[i+1], f"Sheet_{i//2 + 1}")

wb.save("final_output.xlsx")